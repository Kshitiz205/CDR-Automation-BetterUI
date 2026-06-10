from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings, io, os, threading, uuid, time

warnings.filterwarnings("ignore")
app = Flask(__name__)

# ── In-memory job store ───────────────────────────────────────────────────────
jobs = {}   # job_id -> { status, progress, message, result_bytes, stats, error }

# ── CDR ENGINE ────────────────────────────────────────────────────────────────
RAW_COLS = [
    "Subscriber ID","Mobile Number","Session Start","Session End",
    "Online Time(Hr.)","Uploaded MB","Downloaded MB","Total MB",
    "IP Address","MAC Address","BT Site ID","AP Name","Hotspot Name",
    "Circle","Plan Name"
]
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
PVT_FILL  = PatternFill("solid", fgColor="2E75B6")
PVT_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
DATA_FONT = Font(name="Calibri", size=10)

def auto_fit(ws, max_w=45):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best+2, max_w)

def load_mtd(master_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=True, read_only=True)
    mtd_mob = set()
    if "MTD(Mobile No.)" in wb.sheetnames:
        for row in wb["MTD(Mobile No.)"].iter_rows(min_row=2, values_only=True):
            if row[1]: mtd_mob.add(str(row[1]).strip())
    mtd_mac = set()
    if "MTD(MAC)" in wb.sheetnames:
        for row in wb["MTD(MAC)"].iter_rows(min_row=2, values_only=True):
            if row[1]: mtd_mac.add(str(row[1]).strip())
    wb.close()
    return mtd_mob, mtd_mac

def load_and_enrich(input_bytes):
    xl = pd.ExcelFile(io.BytesIO(input_bytes))
    df = pd.read_excel(io.BytesIO(input_bytes), sheet_name=xl.sheet_names[0])
    df.columns = [c.strip() for c in df.columns]
    for col in ["Uploaded MB","Downloaded MB","Total MB"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    cdr_date = df["Session Start"].dropna().dt.date.mode()[0]
    df["Plan Name"] = "Free User"
    df["Mint"] = (
        (df["Session End"]-df["Session Start"])
        .dt.total_seconds().div(60).clip(lower=0).fillna(0).astype(int)
    )
    def to_hms(m):
        h,mn = divmod(int(m),60); return f"{h:02d}:{mn:02d}:00"
    df["Online Time(Hr.)"] = df["Mint"].apply(to_hms)
    df["abc"] = df["Mobile Number"].astype(str)+"&"+df["BT Site ID"].astype(str)
    df["efg"] = df["MAC Address"].astype(str)+"&"+df["BT Site ID"].astype(str)
    df["xyz"] = df["abc"].iloc[::-1].groupby(df["abc"].iloc[::-1]).cumcount().iloc[::-1]+1
    df["hij"] = df["efg"].iloc[::-1].groupby(df["efg"].iloc[::-1]).cumcount().iloc[::-1]+1
    return df, cdr_date

def build_filtered(df, mtd_mob, mtd_mac):
    tu_mask     = ~df.duplicated(subset="abc", keep="last")
    tu_df       = df[tu_mask][RAW_COLS].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    uu_mask     = tu_mask & ~df["abc"].isin(mtd_mob)
    uu_df       = df[uu_mask][RAW_COLS+["abc"]].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    mac_tu_mask = ~df.duplicated(subset="efg", keep="last")
    mac_tu_df   = df[mac_tu_mask][RAW_COLS].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    mac_uu_mask = mac_tu_mask & ~df["efg"].isin(mtd_mac)
    mac_uu_df   = df[mac_uu_mask][RAW_COLS+["efg"]].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    return tu_df, uu_df, mac_tu_df, mac_uu_df

def pivot_cdr(df):
    p = df.groupby("BT Site ID").agg(
        a=("Session Start","count"),b=("Downloaded MB","sum"),
        c=("Uploaded MB","sum"),d=("Total MB","sum")
    ).reset_index()
    p.columns = ["Row Labels","Count of Session Start","Sum of Downloaded MB","Sum of Uploaded MB","Sum of Total MB"]
    return p.sort_values("Row Labels").reset_index(drop=True)

def pivot_mcdr(df, uu_df, tu_df):
    uu_cnt = uu_df.groupby("BT Site ID").size().reset_index(name="UU").rename(columns={"BT Site ID":"Row Labels"})
    tu_cnt = tu_df.groupby("BT Site ID").size().reset_index(name="TU").rename(columns={"BT Site ID":"Row Labels"})
    p = df.groupby("BT Site ID").agg(
        a=("Session Start","count"),b=("Mint","sum"),
        c=("Uploaded MB","sum"),d=("Downloaded MB","sum"),e=("Total MB","sum")
    ).reset_index()
    p.columns = ["Row Labels","Count of Session Start","Sum of Mint","Sum of Uploaded MB","Sum of Downloaded MB","Sum of Total MB"]
    p = p.merge(uu_cnt,on="Row Labels",how="left").merge(tu_cnt,on="Row Labels",how="left")
    p[["UU","TU"]] = p[["UU","TU"]].fillna(0).astype(int)
    return p.sort_values("Row Labels").reset_index(drop=True)

def plan_pivot(src_df):
    p = (src_df.groupby(["BT Site ID","Plan Name"]).size()
         .unstack(fill_value=0).reset_index().rename(columns={"BT Site ID":"Row Labels"}))
    p["Grand Total"] = p.drop(columns="Row Labels").sum(axis=1)
    return p.sort_values("Row Labels").reset_index(drop=True)

def write_pivot_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    ws.append([]); ws.append([])
    ws.append(list(df.columns))
    for c in ws[3]:
        c.fill=PVT_FILL; c.font=PVT_FONT
        c.alignment=Alignment(horizontal="center",vertical="center")
    num_cols = {i+1 for i,col in enumerate(df.columns) if df[col].dtype in ("int64","float64") and col!="Row Labels"}
    for r_idx,row in enumerate(dataframe_to_rows(df,index=False,header=False),4):
        ws.append(row)
        fill=ALT_FILL if r_idx%2==0 else None
        for c_idx in range(1,len(df.columns)+1):
            cell=ws.cell(r_idx,c_idx); cell.font=DATA_FONT
            cell.alignment=Alignment(horizontal="center" if c_idx>1 else "left")
            if fill: cell.fill=fill
            if c_idx in num_cols and isinstance(cell.value,(int,float)):
                cell.number_format="#,##0"
    auto_fit(ws); ws.freeze_panes="A4"

def _style_rows(ws, col_names, start_row=2):
    dt_idx  = {i+1 for i,c in enumerate(col_names) if c in {"Session Start","Session End"}}
    int_idx = {i+1 for i,c in enumerate(col_names) if c in {"Uploaded MB","Downloaded MB","Total MB"}}
    for r in range(start_row, ws.max_row+1):
        fill=ALT_FILL if r%2==0 else None
        for c in dt_idx:
            cell=ws.cell(r,c); cell.font=DATA_FONT
            if fill: cell.fill=fill
            if isinstance(cell.value,datetime):
                cell.number_format="YYYY-MM-DD HH:MM:SS"
                cell.alignment=Alignment(horizontal="center")
        for c in int_idx:
            cell=ws.cell(r,c); cell.font=DATA_FONT
            if fill: cell.fill=fill
            try:
                cell.value=int(float(cell.value)); cell.number_format="#,##0"
                cell.alignment=Alignment(horizontal="right")
            except: pass

def write_data_sheet(wb, name, df):
    ws=wb.create_sheet(name); ws.append(list(df.columns))
    for c in ws[1]:
        c.fill=HDR_FILL; c.font=HDR_FONT
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=20
    for row in dataframe_to_rows(df,index=False,header=False): ws.append(row)
    _style_rows(ws,list(df.columns)); auto_fit(ws); ws.freeze_panes="A2"

def write_cdr_date_sheet(wb, name, df, mtd_mob, mtd_mac):
    ws=wb.create_sheet(name)
    headers=RAW_COLS+["abc","xyz","pqr","Mint","efg","hij","klm"]
    ws.append(headers)
    for c in ws[1]:
        c.fill=HDR_FILL; c.font=HDR_FONT
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=20
    n=len(df); last=n+1
    raw_vals=df[RAW_COLS].values.tolist()
    abc_vals=df["abc"].tolist(); efg_vals=df["efg"].tolist()
    pqr_vals=[k if k in mtd_mob else "#N/A" for k in abc_vals]
    klm_vals=[k if k in mtd_mac else "#N/A" for k in efg_vals]
    for i,rv in enumerate(raw_vals):
        r=i+2
        ws.append(rv+[
            f'=B{r}&"&"&K{r}', f'=COUNTIF(P{r}:P{last+1},P{r})', pqr_vals[i],
            f'=INT((D{r}-C{r})*1440)', f'=J{r}&"&"&K{r}', f'=COUNTIF(T{r}:T{last+1},T{r})', klm_vals[i],
        ])
    _style_rows(ws,headers)
    for r in range(2,n+2):
        fill=ALT_FILL if r%2==0 else None
        for c in range(16,23):
            cell=ws.cell(r,c); cell.font=DATA_FONT
            if fill: cell.fill=fill
            if isinstance(cell.value,int):
                cell.number_format="#,##0"; cell.alignment=Alignment(horizontal="right")
    auto_fit(ws); ws.freeze_panes="A2"

def run_generation(job_id, input_bytes, master_bytes):
    job = jobs[job_id]
    steps = [
        ("Loading MTD lookup tables…",        15),
        ("Parsing session data…",             30),
        ("Applying TU / UU / MAC filters…",   48),
        ("Building pivot tables…",            63),
        ("Writing 11 Excel sheets…",          82),
        ("Finalising output file…",           95),
    ]
    try:
        job.update({"step": steps[0][0], "progress": 5})
        mtd_mob, mtd_mac = load_mtd(master_bytes)
        job.update({"step": steps[0][0], "progress": steps[0][1]})

        job.update({"step": steps[1][0], "progress": steps[0][1]})
        df, cdr_date = load_and_enrich(input_bytes)
        job.update({"step": steps[1][0], "progress": steps[1][1]})

        job.update({"step": steps[2][0], "progress": steps[1][1]})
        tu_df, uu_df, mac_tu_df, mac_uu_df = build_filtered(df, mtd_mob, mtd_mac)
        job.update({"step": steps[2][0], "progress": steps[2][1]})

        job.update({"step": steps[3][0], "progress": steps[2][1]})
        p_cdr      = pivot_cdr(df)
        p_mcdr     = pivot_mcdr(df, uu_df, tu_df)
        uup_df     = plan_pivot(uu_df)
        tup_df     = plan_pivot(tu_df)
        mac_uup_df = plan_pivot(mac_uu_df)
        mac_tup_df = plan_pivot(mac_tu_df)
        job.update({"step": steps[3][0], "progress": steps[3][1]})

        job.update({"step": steps[4][0], "progress": steps[3][1]})
        wb = Workbook(); wb.remove(wb.active)
        date_sheet = f"CDR{cdr_date.strftime('%Y-%m-%d')}"
        write_pivot_sheet(wb,"CDR",p_cdr)
        write_pivot_sheet(wb,"M.CDR",p_mcdr)
        write_cdr_date_sheet(wb,date_sheet,df,mtd_mob,mtd_mac)
        write_pivot_sheet(wb,"MAC UUP",mac_uup_df)
        write_data_sheet(wb,"MAC UU",mac_uu_df)
        write_pivot_sheet(wb,"MAC TUP",mac_tup_df)
        write_data_sheet(wb,"MAC TU",mac_tu_df)
        write_pivot_sheet(wb,"UUP",uup_df)
        write_data_sheet(wb,"UU",uu_df)
        write_pivot_sheet(wb,"TUP",tup_df)
        write_data_sheet(wb,"TU",tu_df)
        job.update({"step": steps[4][0], "progress": steps[4][1]})

        job.update({"step": steps[5][0], "progress": steps[4][1]})
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        job.update({
            "step": "Done!", "progress": 100, "status": "done",
            "result_bytes": buf.getvalue(),
            "stats": {
                "date": str(cdr_date),
                "sessions": len(df),
                "bts": int(df["BT Site ID"].nunique()),
                "tu": len(tu_df), "uu": len(uu_df),
                "mac_tu": len(mac_tu_df), "mac_uu": len(mac_uu_df),
                "sheets": len(wb.sheetnames)
            }
        })
    except Exception as e:
        job.update({"status": "error", "error": str(e), "progress": 0})

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/generate", methods=["POST"])
def generate():
    if "input_file" not in request.files or "master_file" not in request.files:
        return jsonify({"error": "Both files required"}), 400
    input_bytes  = request.files["input_file"].read()
    master_bytes = request.files["master_file"].read()
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "running", "progress": 0, "step": "Starting…", "result_bytes": None, "stats": None, "error": None}
    t = threading.Thread(target=run_generation, args=(job_id, input_bytes, master_bytes))
    t.daemon = True
    t.start()
    return jsonify({"job_id": job_id})

@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job: return jsonify({"error": "Not found"}), 404
    return jsonify({
        "status":   job["status"],
        "progress": job["progress"],
        "step":     job["step"],
        "stats":    job["stats"],
        "error":    job["error"]
    })

@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("result_bytes"):
        return jsonify({"error": "Not ready"}), 404
    fname = request.args.get("filename", "CDR_Final_Output.xlsx")
    if not fname.endswith(".xlsx"): fname += ".xlsx"
    return send_file(
        io.BytesIO(job["result_bytes"]),
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(debug=False, port=5000)
